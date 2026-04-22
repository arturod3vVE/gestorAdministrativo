from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from datetime import datetime, timedelta
from django.contrib.auth.models import User
from django.db.models import Sum, Q, Case, When, Value
from django.http import HttpResponseForbidden
from django.db import transaction
from django.contrib import messages
from django.db import transaction
from django.utils import timezone
from django.http import HttpResponse
from django.core.paginator import Paginator
from decimal import Decimal, ROUND_HALF_UP
import json
import csv
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum, Q, Case, When, Value, DecimalField, Count
from django.db.models.functions import Coalesce
from django.utils import timezone
from datetime import timedelta
from .models import *
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.db.models.functions import TruncMonth
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
import json

# Importamos los modelos y formularios actualizados
from .models import Socio, ConfigSistema, Pago, MetodoPago, ConceptoCobro, GastoGeneral, AvisoCobro, ItemAviso, CierreMes
from .forms import SocioForm, RegistroSocioForm, ConceptoForm, SocioEditForm, ConfigTelefonoForm
from .utils import obtener_tasa_bcv

# --- FUNCIONES AUXILIARES ---

def es_admin(user):
    return user.is_authenticated and user.is_staff

@login_required
def redireccionar_por_rol(request):
    if request.user.is_staff:
        return redirect('dashboard')

    # Verificamos que sea un socio antes de mandarlo al portal
    if hasattr(request.user, 'socio'):
        return redirect('portal_socio')

    # Si no es admin ni socio (un usuario raro), lo sacamos
    messages.error(request, "Tu cuenta no tiene un perfil de socio asignado.")
    logout(request)
    return redirect('login')

from django.contrib.auth import logout
def salir_del_sistema(request):
    logout(request)
    return redirect('login')

# --- VISTAS DEL SOCIO ---

@login_required
def portal_socio(request):
    if not hasattr(request.user, 'socio'):
        return redirect('dashboard')

    socio_actual = request.user.socio
    hoy = timezone.now()

    # Traemos los avisos con sus detalles y pagos de un solo golpe a la memoria
    avisos = AvisoCobro.objects.filter(socio=socio_actual).prefetch_related('detalles', 'pagos').order_by('-anio', '-mes')

    config = ConfigSistema.objects.first()
    # Mantenemos la tasa como Decimal
    tasa_actual = config.tasa_bcv if config else Decimal('36.0000')

    # Constantes útiles para operaciones
    CERO = Decimal('0.00')
    CENTAVO = Decimal('0.01')
    total_deuda_usd = CERO

    for aviso in avisos:
        # Matemática rápida en memoria usando Decimal puro
        # Usamos generator expressions y le pasamos 'CERO' como valor inicial del sum()
        facturado = sum((d.monto_dolares for d in aviso.detalles.all()), CERO)
        pagado_aprobado = sum((p.monto_dolares for p in aviso.pagos.all() if p.estado == 'APROBADO'), CERO)
        pagado_revision = sum((p.monto_dolares for p in aviso.pagos.all() if p.estado == 'REVISION'), CERO)
        
        saldo_real = facturado - pagado_aprobado
        saldo_por_cubrir = facturado - pagado_aprobado - pagado_revision

        # Atributos limpios para el HTML
        aviso.monto_total_usd = facturado
        aviso.saldo_pendiente = saldo_real
        aviso.en_revision = pagado_revision > CERO
        
        aviso.esta_pagado = (aviso.estado == 'PAGADO') or (saldo_real <= CERO)
        aviso.puede_pagar = not aviso.esta_pagado and (saldo_por_cubrir >= CENTAVO)

        if saldo_real > CERO:
            total_deuda_usd += saldo_real

    # Cálculo seguro de la deuda total en bolívares con redondeo comercial exacto
    total_deuda_bs = (total_deuda_usd * tasa_actual).quantize(CENTAVO, rounding=ROUND_HALF_UP)

    context = {
        'socio': socio_actual,
        'avisos': avisos,
        'total_deuda': total_deuda_usd,
        'total_deuda_bs': total_deuda_bs,
        'tasa': tasa_actual,
        'fecha_actual': hoy,
        'mes_actual': hoy.month,
        'anio_actual': hoy.year,
    }
    return render(request, 'core/portal_socio.html', context)


@login_required
def mis_pagos(request):
    # Capturamos filtros del navegador
    mes_filtro = request.GET.get('mes')
    anio_filtro = request.GET.get('anio')

    # Base de la consulta: solo los pagos del usuario logueado
    pagos = Pago.objects.filter(aviso__socio__user=request.user).select_related('aviso', 'metodo')

    # Aplicamos filtros si existen
    if mes_filtro:
        pagos = pagos.filter(aviso__mes=mes_filtro)
    if anio_filtro:
        pagos = pagos.filter(aviso__anio=anio_filtro)

    # Obtenemos años únicos donde el socio tiene pagos para el selector
    años_disponibles = Pago.objects.filter(
        aviso__socio__user=request.user
    ).values_list('aviso__anio', flat=True).distinct().order_by('-aviso__anio')

    return render(request, 'core/mis_pagos.html', {
        'pagos': pagos,
        'años': años_disponibles,
        'mes_actual': mes_filtro,
        'anio_actual': anio_filtro,
        'meses_nombres': [
            (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
            (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
            (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
        ]
    })


@login_required
def reportar_pago(request, aviso_id):
    # Seguridad: Validar que sea un socio y buscar el aviso
    socio_actual = getattr(request.user, 'socio', None)
    if not socio_actual:
        return redirect('dashboard')
        
    aviso = get_object_or_404(AvisoCobro, pk=aviso_id, socio=socio_actual)
    
    config = ConfigSistema.objects.first()
    # REEMPLAZO 1: Mantenemos la tasa como Decimal
    tasa = config.tasa_bcv if config else Decimal('0.0000')

    if aviso.estado == 'PAGADO':
        messages.success(request, "¡Este aviso ya está solvente!")
        return redirect('portal_socio')

    # Cálculos dinámicos del Aviso de Cobro
    # REEMPLAZO 2: Los fallbacks "0" cambian a Decimal('0.00')
    total_facturado_usd = aviso.detalles.aggregate(total=Sum('monto_dolares'))['total'] or Decimal('0.00')
    total_aprobado_usd = aviso.pagos.filter(estado='APROBADO').aggregate(total=Sum('monto_dolares'))['total'] or Decimal('0.00')
    total_pendiente_usd = aviso.pagos.filter(estado='REVISION').aggregate(total=Sum('monto_dolares'))['total'] or Decimal('0.00')

    # Lo que realmente puede pagar el socio ahorita
    # REEMPLAZO 3: Eliminados los float(), operamos directamente con los Decimal
    saldo_pagable_usd = total_facturado_usd - total_aprobado_usd - total_pendiente_usd

    # ¡BARRERA DE SEGURIDAD INDIVIDUAL!
    # REEMPLAZO 4: Decimal exacto
    if saldo_pagable_usd <= Decimal('0.01'):
        messages.info(request, "🛡️ El pago total de este aviso ya está en revisión por la administración.")
        return redirect('portal_socio')

    # Si pasa la validación, calculamos el máximo en Bs.
    # REEMPLAZO 5: Forzamos que este cálculo de vista también tenga 2 decimales
    saldo_real_bs = (saldo_pagable_usd * tasa).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    metodos_activos = MetodoPago.objects.filter(activo=True).order_by('tipo')

    if request.method == 'POST':
        monto_bs_str = request.POST.get('monto_bs', '0').replace(',', '.')
        monto_usd_str = request.POST.get('monto_usd', '0').replace(',', '.')
        referencia = request.POST.get('referencia')
        comprobante = request.FILES.get('comprobante')
        metodo_id = request.POST.get('metodo')

        if not metodo_id:
                messages.error(request, "⚠️ Error: Debes seleccionar un banco.")
        elif not referencia or not comprobante:
                messages.error(request, "⚠️ Error: Debes ingresar el número de referencia y adjuntar el comprobante.")
        else:
            try:
                banco_obj = MetodoPago.objects.get(pk=metodo_id)
                    
                # MAGIA MULTIMONEDA: Calculamos dependiendo de la configuración del banco
                # REEMPLAZO 6: Casteos a Decimal y reemplazo de round() por quantize()
                if banco_obj.moneda == 'BS':
                    monto_bs = Decimal(monto_bs_str)
                    monto_usd = (monto_bs / tasa).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                else:
                    monto_usd = Decimal(monto_usd_str)
                    monto_bs = (monto_usd * tasa).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

                # 2. VALIDACIÓN ESTRICTA EN DÓLARES (La moneda base)
                # REEMPLAZO 7: Condicionales en Decimal
                if monto_usd > (saldo_pagable_usd + Decimal('0.01')):
                    messages.error(request, f"❌ Error: El monto máximo permitido para este aviso es ${saldo_pagable_usd:.2f}.")
                elif monto_usd <= Decimal('0.00'):
                    messages.error(request, "❌ Error: El monto debe ser mayor a cero.")
                else:
                    with transaction.atomic():
                        Pago.objects.create(
                            aviso=aviso,
                            metodo=banco_obj,
                            referencia=referencia,
                            comprobante=comprobante,
                            tasa_bcv_usada=tasa,
                            monto_bolivares=monto_bs,
                            monto_dolares=monto_usd,
                            estado='REVISION'
                        )
                        aviso.estado = 'REVISION'
                        aviso.save()

                    messages.success(request, "✅ Pago reportado con éxito.")
                    return redirect('portal_socio')
                        
            except MetodoPago.DoesNotExist:
                    messages.error(request, "⚠️ Error: El método de pago no es válido.")
            # REEMPLAZO 8: Decimal puede arrojar InvalidOperation además de ValueError
            except (ValueError, TypeError, ZeroDivisionError, Exception) as e:
                    messages.error(request, "⚠️ Error: Monto o tasa inválidos. Usa solo números y puntos/comas.")

    # Contexto totalmente transparente y coincidente con tu HTML
    context = {
        'aviso': aviso,
        'metodos_pago': metodos_activos,
        'tasa': tasa,
        'saldo_bs': saldo_real_bs,
        'saldo_usd': saldo_pagable_usd,
    }
    return render(request, 'core/form_pago.html', context)

@login_required
def recibo_mensual(request, mes, anio):
    socio = request.user.socio
    config = ConfigSistema.objects.first()
    tasa = config.tasa_bcv if config else Decimal('0.0000')

    aviso = AvisoCobro.objects.filter(socio=socio, mes=mes, anio=anio).first()
    detalles = aviso.detalles.all() if aviso else []
    
    # Sumatoria pura en Decimal
    total_usd = sum((c.monto_dolares for c in detalles), Decimal('0.00'))
    total_bs = (total_usd * tasa).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    return render(request, 'core/recibo_mensual.html', {
        'aviso': aviso, 
        'cargos': detalles, 
        'total_usd': total_usd,
        'total_bs': total_bs, 
        'tasa': tasa, 
        'mes_nombre': meses[int(mes)-1],
        'anio': anio, 
        'config': config
    })


def recibo_print(request, aviso_id):
    aviso = get_object_or_404(AvisoCobro, pk=aviso_id)
    config = ConfigSistema.objects.first()
    
    # 1. Aseguramos que la tasa sea un Decimal válido incluso si no hay config
    tasa = config.tasa_bcv if config else Decimal('0.0000')
    
    # 2. Le pasamos Decimal('0.00') al sum() para que no empiece a sumar desde un entero '0'
    total_usd = sum((d.monto_dolares for d in aviso.detalles.all()), Decimal('0.00'))
    
    # 3. Multiplicación pura en Decimal y redondeo a 2 dígitos para el recibo
    monto_bs = (total_usd * tasa).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    
    return render(request, 'core/recibo_print.html', {
        'aviso': aviso, 
        'tasa': tasa, 
        'monto_usd': total_usd, 
        'monto_bs': monto_bs
    })


@login_required
def previsualizar_aviso(request, aviso_id):
    # Buscamos el aviso
    aviso = get_object_or_404(AvisoCobro, id=aviso_id)
    
    # SEGURIDAD: Si no es admin y el aviso no le pertenece, bloqueamos el acceso
    if not request.user.is_staff and aviso.socio.user != request.user:
        messages.error(request, "No tienes permiso para ver este documento.")
        return redirect('portal_socio')

    config = ConfigSistema.objects.first()
    
    # 1. Tasa pura en Decimal
    tasa_calculo = config.tasa_bcv if config else Decimal('0.0000')
    mensaje_tasa = f"Calculado a la tasa oficial del día: Bs. {tasa_calculo}"

    items_con_bs = []
    
    # 2. Inicializamos los totales en Decimal exacto, no en enteros (0)
    total_bs = Decimal('0.00')
    total_usd = Decimal('0.00')
    
    for item in aviso.detalles.all():
        # 3. Multiplicación Decimal pura y redondeo comercial ítem por ítem
        monto_bs = (item.monto_dolares * tasa_calculo).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        items_con_bs.append({
            'descripcion': item.descripcion, 
            'monto_bs': monto_bs, 
            'monto_usd': item.monto_dolares
        })
        
        # 4. Acumulamos sumando solo Decimals
        total_bs += monto_bs
        total_usd += item.monto_dolares

    return render(request, 'core/aviso_cobro_pdf.html', {
        'aviso': aviso, 
        'items': items_con_bs, 
        'total_usd': total_usd,
        'total_bs': total_bs, 
        'tasa': tasa_calculo, 
        'mensaje_tasa': mensaje_tasa, 
        'config': config
    })

# --- VISTAS ADMINISTRATIVAS ---

@user_passes_test(es_admin)
def dashboard(request):
    hoy = timezone.now()
    
    # ==========================================================
    # 1. CONFIGURACIÓN DE TASA BCV (Tu lógica original recuperada)
    # ==========================================================
    config, _ = ConfigSistema.objects.get_or_create(id=1, defaults={'tasa_bcv': Decimal('400.00')})

    if request.GET.get('actualizar_bcv'):
        try:
            nueva_tasa = obtener_tasa_bcv() # Tu función de web scraping/API
            if nueva_tasa:
                config.tasa_bcv = Decimal(str(nueva_tasa))
                config.save(user_audit=request.user)
                messages.success(request, f"✅ Tasa actualizada desde el BCV: {nueva_tasa}")
        except Exception as e:
            messages.error(request, f"❌ Error al conectar con el servicio del BCV.")
        return redirect('dashboard')

    if request.method == 'POST' and 'tasa_manual' in request.POST:
        try:
            config.tasa_bcv = Decimal(request.POST.get('tasa_manual').replace(',', '.'))
            config.save(user_audit=request.user)
            messages.success(request, "✅ Tasa manual establecida.")
        except ValueError:
            messages.error(request, "❌ Formato de tasa inválido.")
        return redirect('dashboard')

    # ==========================================================
    # 2. SISTEMA DE FILTROS (AÑO Y MES)
    # ==========================================================
    anios_db = AvisoCobro.objects.values_list('anio', flat=True).distinct().order_by('-anio')
    anios_disponibles = list(anios_db) if anios_db else [hoy.year]
    
    anio_sel = int(request.GET.get('anio', hoy.year))
    mes_sel = int(request.GET.get('mes', 0)) # 0 significa "Todos los meses"

    filtros = Q(anio=anio_sel)
    if mes_sel != 0:
        filtros &= Q(mes=mes_sel)

    avisos_filtrados = AvisoCobro.objects.filter(filtros)

    # ==========================================================
    # 3. CÁLCULO DE MÉTRICAS FILTRADAS
    # ==========================================================
    # Deuda Pendiente
    total_usd = sum(a.saldo_pendiente_usd for a in avisos_filtrados.exclude(estado='PAGADO'))
    total_bs = total_usd * config.tasa_bcv

    # Solvencia
    conteo_socios = Socio.objects.filter(activo=True).count()
    if mes_sel != 0:
        solventes = Socio.objects.filter(activo=True).exclude(
            avisos__anio=anio_sel, 
            avisos__mes=mes_sel,
            avisos__estado__in=['PENDIENTE', 'PARCIAL', 'REVISION']
        ).distinct().count()
    else:
        solventes = Socio.objects.filter(activo=True).exclude(
            avisos__anio=anio_sel, 
            avisos__estado__in=['PENDIENTE', 'PARCIAL', 'REVISION']
        ).distinct().count()
        
    porcentaje_solvencia = int((solventes / conteo_socios * 100)) if conteo_socios > 0 else 0

    # Recuperado (Monto de pagos aprobados en el periodo seleccionado)
    recaudado_query = Pago.objects.filter(aviso__anio=anio_sel, estado='APROBADO')
    if mes_sel != 0:
        recaudado_query = recaudado_query.filter(aviso__mes=mes_sel)
        
    total_recaudado = recaudado_query.aggregate(total=Sum('monto_dolares'))['total'] or Decimal('0.00')

    # ==========================================================
    # 4. GRÁFICO (Siempre muestra todo el año para ver tendencia)
    # ==========================================================
    recaudacion_chart = (
        Pago.objects.filter(aviso__anio=anio_sel, estado='APROBADO')
        .annotate(mes_pago=TruncMonth('fecha_reporte'))
        .values('mes_pago')
        .annotate(total=Sum('monto_dolares'))
        .order_by('mes_pago')
    )
    
    meses_es = {1:'Ene', 2:'Feb', 3:'Mar', 4:'Abr', 5:'May', 6:'Jun', 7:'Jul', 8:'Ago', 9:'Sep', 10:'Oct', 11:'Nov', 12:'Dic'}
    
    chart_labels = []
    chart_values = []
    for r in recaudacion_chart:
        if r['mes_pago']:
            chart_labels.append(meses_es[r['mes_pago'].month])
            chart_values.append(float(r['total']))

    # Tu fallback original si no hay datos
    if not chart_labels:
        chart_labels = [meses_es[hoy.month]]
        chart_values = [0.0]

    # ==========================================================
    # 5. TABLAS Y CONTEXTO
    # ==========================================================
    avisos_recientes = avisos_filtrados.order_by(
        Case(When(estado='REVISION', then=Value(0)), default=Value(1)),
        '-fecha_emision'
    )[:10]

    meses_nombres = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'), (5, 'Mayo'), (6, 'Junio'),
        (7, 'Julio'), (8, 'Agosto'), (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]

    context = {
        'config': config,
        'fecha_actual': hoy,
        'mes_cerrado': CierreMes.objects.filter(mes=hoy.month, anio=hoy.year).exists(),
        
        # Filtros
        'anio_actual': anio_sel,
        'mes_actual': mes_sel,
        'anios_disponibles': anios_disponibles,
        'meses_nombres': meses_nombres,
        
        # Métricas
        'total_usd': total_usd,
        'total_bs': total_bs,
        'porcentaje_solvencia': porcentaje_solvencia,
        'total_auditoria': total_recaudado,
        
        # Gráfico y Tabla
        'chart_labels': json.dumps(chart_labels),
        'chart_data': json.dumps(chart_values),
        'avisos_recientes': avisos_recientes,
    }
    return render(request, 'core/dashboard.html', context)

@login_required
def lista_transparencia(request):
    socios_qs = Socio.objects.annotate(
        total_facturado=Coalesce(
            Sum('avisos__detalles__monto_dolares', filter=~Q(avisos__estado='PAGADO')), 
            Value(Decimal('0.00')), 
            output_field=DecimalField()
        ),
        total_pagado=Coalesce(
            Sum('avisos__pagos__monto_dolares', filter=Q(avisos__pagos__estado='APROBADO') & ~Q(avisos__estado='PAGADO')), 
            Value(Decimal('0.00')), 
            output_field=DecimalField()
        )
    ).annotate(
        # La base de datos hace la resta por nosotros
        deuda_total=F('total_facturado') - F('total_pagado')
    ).order_by('unidad')

    # 2. Capturamos filtros
    query = request.GET.get('q')
    status_filter = request.GET.get('status')
    debt_filter = request.GET.get('debt')

    # 3. Aplicamos filtros directamente en la BD (Súper eficiente)
    if query:
        socios_qs = socios_qs.filter(Q(nombre__icontains=query) | Q(unidad__icontains=query) | Q(cedula__icontains=query))

    if status_filter == 'activo':
        socios_qs = socios_qs.filter(activo=True)
    elif status_filter == 'inactivo':
        socios_qs = socios_qs.filter(activo=False)

    if debt_filter == 'con_deuda':
        socios_qs = socios_qs.filter(deuda_total__gt=Decimal('0.01'))
    elif debt_filter == 'solvente':
        socios_qs = socios_qs.filter(deuda_total__lte=Decimal('0.01'))

    # 4. Paginación
    paginator = Paginator(socios_qs, 10)
    page_number = request.GET.get('page')
    socios_paginados = paginator.get_page(page_number)

    # 5. Totales Globales (Directo en la BD)
    totales = socios_qs.aggregate(
        gran_total_deuda=Sum('deuda_total')
    )
    total_global_deuda = totales['gran_total_deuda'] or Decimal('0.00')
    
    socios_solventes_count = socios_qs.filter(deuda_total__lte=Decimal('0.01')).count()

    return render(request, 'core/transparencia.html', {
        'socios': socios_paginados, 
        'total_global_deuda': total_global_deuda,
        'socios_solventes_count': socios_solventes_count, 
        'filtros': request.GET.dict()
    })


@user_passes_test(es_admin)
def registrar_socio_completo(request):
    if request.method == 'POST':
        form = RegistroSocioForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # FORMATEO: Letra + 9 dígitos (relleno con ceros)
                    cedula = f"{form.cleaned_data['nacionalidad']}{form.cleaned_data['cedula_numero'].zfill(9)}"

                    # FORMATEO: 58 + Prefijo (sin el 0) + 7 dígitos
                    prefijo = form.cleaned_data['prefijo_tlf']
                    telefono = f"58{prefijo[1:]}{form.cleaned_data['cuerpo_tlf']}"

                    user = User.objects.create_user(
                        username=form.cleaned_data['username'].lower(),
                        password=form.cleaned_data['password'],
                        first_name=form.cleaned_data['nombre']
                    )

                    socio = form.save(commit=False)
                    socio.user = user
                    socio.cedula = cedula
                    socio.telefono = telefono
                    socio.save()

                # Guardamos para el botón de WhatsApp
                request.session['registro_exitoso'] = {
                    'nombre': socio.nombre,
                    'usuario': user.username,
                    'clave': form.cleaned_data['password'],
                    'telefono': telefono
                }
                messages.success(request, f"Socio {cedula} creado con éxito.")
                return redirect('registrar_socio_completo')
            except Exception as e:
                messages.error(request, f"Error: {e}")
    else:
        form = RegistroSocioForm()

    datos_whatsapp = request.session.pop('registro_exitoso', None)
    return render(request, 'core/registrar_socio.html', {'form': form, 'datos_whatsapp': datos_whatsapp})


@login_required
@user_passes_test(es_admin)
def generar_cobros_masivos(request):
    conceptos = ConceptoCobro.objects.all().order_by('nombre')

    if request.method == 'POST':
        concepto_id = request.POST.get('concepto_id')
        monto_manual = request.POST.get('monto_usd')

        if not concepto_id:
            messages.error(request, "Selecciona un concepto.")
            return redirect('generar_cobros_masivos')

        concepto_obj = get_object_or_404(ConceptoCobro, pk=concepto_id)
        
        # Procesamos el monto para el log y la lógica
        precio_forzado = None
        if monto_manual and monto_manual.strip():
            try:
                # Usamos Decimal para evitar los problemas de redondeo de float que vimos antes
                precio_forzado = Decimal(monto_manual.replace(',', '.'))
            except ValueError:
                pass

        socios = Socio.objects.filter(activo=True)
        count = 0
        
        try:
            with transaction.atomic():
                for socio in socios:
                    monto = precio_forzado if precio_forzado is not None else (concepto_obj.monto_chofer if socio.tiene_avance else concepto_obj.monto_sugerido)
                    
                    if monto == 0 and socio.tiene_avance:
                        monto = concepto_obj.monto_sugerido
                        
                    cargar_deuda_a_socio(socio, concepto_obj.nombre, monto)
                    count += 1

            # --- REGISTRO EN BITÁCORA ---
            detalle_monto = f"${precio_forzado} (Manual)" if precio_forzado else "Monto por defecto del concepto"
            
            BitacoraAuditoria.objects.create(
                admin=request.user,
                accion='CARGO_MASIVO', # Asegúrate de añadir esta acción a tus CHOICES en models.py
                modulo='Cobranzas',
                descripcion=(
                    f"Generación masiva de cobros: '{concepto_obj.nombre}'. "
                    f"Monto aplicado: {detalle_monto}. "
                    f"Total socios afectados: {count}."
                ),
                ip_address=get_client_ip(request)
            )

            messages.success(request, f"✅ Se aplicó el cobro de '{concepto_obj.nombre}' a {count} facturas.")
            return redirect('modulo_cobranza')
            
        except Exception as e:
            messages.error(request, f"Error en el proceso: {e}")

    return render(request, 'core/generar_masivo.html', {'conceptos': conceptos})

@login_required
@user_passes_test(es_admin)
def modulo_cobranza(request):
    # Anotamos los cálculos directamente en la base de datos
    socios_pendientes = Socio.objects.filter(activo=True).annotate(
        total_facturado=Coalesce(
            Sum('avisos__detalles__monto_dolares', filter=~Q(avisos__estado='PAGADO')), 
            Value(Decimal('0.00')), output_field=DecimalField()
        ),
        total_aprobado=Coalesce(
            Sum('avisos__pagos__monto_dolares', filter=Q(avisos__pagos__estado='APROBADO') & ~Q(avisos__estado='PAGADO')), 
            Value(Decimal('0.00')), output_field=DecimalField()
        ),
        pagos_nuevos=Count('avisos__pagos', filter=Q(avisos__pagos__estado='REVISION')),
        # Usamos distinct=True para que los JOINs no dupliquen las facturas
        cantidad_facturas=Count('avisos', filter=~Q(avisos__estado='PAGADO'), distinct=True)
    ).annotate(
        total_pendiente=F('total_facturado') - F('total_aprobado')
    ).filter(
        # Solo traemos a los que deben dinero o tienen pagos por revisar
        Q(total_pendiente__gt=Decimal('0.00')) | Q(pagos_nuevos__gt=0)
    ).order_by('unidad')

    # Calcular totales globales en una sola consulta
    totales_globales = socios_pendientes.aggregate(
        suma_deuda=Sum('total_pendiente'),
        suma_revisiones=Sum('pagos_nuevos')
    )
    
    total_global_usd = totales_globales['suma_deuda'] or Decimal('0.00')
    pagos_por_aprobar_total = totales_globales['suma_revisiones'] or 0

    return render(request, 'core/cobranza.html', {
        'socios': socios_pendientes,
        'total_global': total_global_usd,
        'pagos_por_aprobar_total': pagos_por_aprobar_total
    })

@login_required
@user_passes_test(es_admin)
def aprobar_rechazar_pago(request, pago_id, accion):
    # Obtenemos el pago inicial para tener la referencia y el socio
    pago_base = get_object_or_404(Pago, pk=pago_id)
    socio = pago_base.aviso.socio
    referencia = pago_base.referencia

    with transaction.atomic():
        # Buscamos todos los registros que comparten la misma referencia de ese socio
        pagos_hermanos = Pago.objects.filter(
            aviso__socio=socio, 
            referencia=referencia, 
            estado='REVISION'
        ).select_related('aviso')
        
        cantidad = pagos_hermanos.count()
        
        if cantidad == 0:
            messages.warning(request, "Este pago ya fue procesado o no existe.")
            return redirect('gestionar_socio_cobranza', socio_id=socio.id)

        for pago in pagos_hermanos:
            aviso = pago.aviso
            
            if accion == 'aprobar':
                pago.estado = 'APROBADO'
                pago.save()
            elif accion == 'rechazar':
                pago.estado = 'RECHAZADO'
                pago.save()
                
            total_deuda = aviso.total_facturado_usd
            pagado_hasta_ahora = aviso.total_pagado_usd
            
            # ELIMINADO EL MARGEN DE ERROR: Ahora la comparación es exacta.
            if pagado_hasta_ahora >= total_deuda:
                aviso.estado = 'PAGADO'
            elif pagado_hasta_ahora > Decimal('0.00'):
                aviso.estado = 'PARCIAL'
            else:
                aviso.estado = 'PENDIENTE'
            
            aviso.save()

    BitacoraAuditoria.objects.create(
        admin=request.user,
        accion='APROBAR' if accion == 'aprobar' else 'RECHAZAR',
        modulo='Cobranzas',
        descripcion=f"Se {'aprobaron' if accion == 'aprobar' else 'rechazaron'} {cantidad} pagos de la referencia {referencia} del socio {socio.unidad} - {socio.nombre}.",
        ip_address=get_client_ip(request)
    )

    # Mensajes finales
    if accion == 'aprobar':
        messages.success(request, f"✅ ¡Éxito! Se aprobaron {cantidad} registros de la ref. {referencia}.")
    else:
        messages.error(request, f"❌ Se rechazaron {cantidad} registros de la ref. {referencia}.")

    return redirect('gestionar_socio_cobranza', socio_id=socio.id)

@user_passes_test(es_admin)
def historial_pagos(request):
    pagos = Pago.objects.select_related('deuda__socio').order_by('-fecha_reporte')
    # Sumar solo aprobados
    total_usd = sum(p.monto_pagado_usd for p in pagos if p.estado == 'APROBADO')
    total_bs = sum(p.monto_pagado_bs for p in pagos if p.estado == 'APROBADO' and p.monto_pagado_bs)

    return render(request, 'core/historial_pagos.html', {
        'pagos': pagos,
        'total_usd': total_usd,
        'total_bs': total_bs
    })


@user_passes_test(es_admin)
def reporte_auditoria(request):
    items_historico = ItemAviso.objects.filter(
        Q(descripcion__icontains='Auditoria')
    ).select_related('aviso').order_by('-aviso__anio', '-aviso__mes')

    recuperado = 0
    pendiente = 0
    for item in items_historico:
        if item.aviso.estado == 'PAGADO':
            recuperado += float(item.monto_dolares)
        else:
            pendiente += float(item.monto_dolares)

    return render(request, 'core/auditoria.html', {
        'historico': items_historico, 'total_recuperado': recuperado, 'total_pendiente': pendiente
    })


# --- CONFIGURACIONES ---

@user_passes_test(es_admin)
def gestionar_conceptos(request):
    if request.method == 'POST':
        form = ConceptoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Concepto creado.")
            return redirect('gestionar_conceptos')
    else:
        form = ConceptoForm()

    conceptos = ConceptoCobro.objects.all().order_by('nombre')
    return render(request, 'core/conceptos_lista.html', {'form': form, 'conceptos': conceptos})


@user_passes_test(es_admin)
def eliminar_concepto(request, id):
    c = get_object_or_404(ConceptoCobro, pk=id)
    c.delete()
    messages.warning(request, "Concepto eliminado.")
    return redirect('gestionar_conceptos')


@login_required
@user_passes_test(lambda u: u.is_staff)
def configuracion_pagos(request):
    if request.method == 'POST':
        try:
            # Capturamos los datos del POST
            tipo = request.POST.get('tipo')
            moneda = request.POST.get('moneda') # <--- El nuevo campo
            banco = request.POST.get('banco')
            beneficiario = request.POST.get('beneficiario')
            cedula = request.POST.get('cedula')
            
            # Campos opcionales según el tipo
            telefono = request.POST.get('telefono') if tipo == 'PAGO_MOVIL' else None
            cuenta = request.POST.get('cuenta') if tipo == 'TRANSFERENCIA' else None
            correo = request.POST.get('correo') if tipo == 'ZELLE' else None

            # Creamos el registro en la base de datos
            MetodoPago.objects.create(
                tipo=tipo,
                moneda=moneda, # <--- Guardamos la moneda
                nombre_banco=banco,
                titular=beneficiario,
                cedula_rif=cedula,
                telefono=telefono,
                numero_cuenta=cuenta,
                correo=correo,
                activo=True # Por defecto se crea activo
            )
            
            messages.success(request, f"Método {tipo} en {moneda} agregado correctamente.")
            
        except Exception as e:
            messages.error(request, f"Error al guardar el método: {e}")
            
        return redirect('configuracion_pagos')

    # Obtenemos todos los métodos para la tabla, ordenados por moneda y tipo
    metodos = MetodoPago.objects.all().order_by('moneda', 'tipo')
    return render(request, 'core/config_pagos.html', {'metodos': metodos})


@login_required
@user_passes_test(es_admin)
def toggle_metodo_pago(request, metodo_id):
    m = get_object_or_404(MetodoPago, pk=metodo_id)
    
    # 1. Togleamos el estado
    m.activo = not m.activo
    m.save()
    
    # 2. Preparamos el texto según el nuevo estado
    estado_texto = "ACTIVÓ" if m.activo else "DESACTIVÓ"
    
    # 3. REGISTRO EN BITÁCORA
    BitacoraAuditoria.objects.create(
        admin=request.user,
        accion='ACTUALIZAR_METODO',
        modulo='Configuración',
        descripcion=f"Se {estado_texto} el método de pago: {m.get_tipo_display()} - {m.nombre_banco} (Titular: {m.titular})",
        ip_address=get_client_ip(request)
    )
    
    # Mensaje visual para el admin
    alert_type = "success" if m.activo else "warning"
    messages.add_message(request, messages.SUCCESS if m.activo else messages.WARNING, 
                         f"El método {m.nombre_banco} ahora está {'Activo' if m.activo else 'Inactivo'}.")
    
    return redirect('configuracion_pagos')

@user_passes_test(es_admin)
@login_required
@user_passes_test(es_admin)
def editar_socio(request, socio_id):
    socio = get_object_or_404(Socio, id=socio_id)
    
    # Capturamos datos actuales antes de la edición para el log
    nombre_previo = socio.nombre
    unidad_previa = socio.unidad

    if request.method == 'POST':
        form = SocioEditForm(request.POST, instance=socio)
        if form.is_valid():
            s = form.save(commit=False)
            
            # RE-FORMATEO ESTRICTO (Tu lógica original)
            s.cedula = f"{form.cleaned_data['nacionalidad']}{form.cleaned_data['cedula_numero'].zfill(9)}"
            pref = form.cleaned_data['prefijo_tlf']
            s.telefono = f"58{pref[1:]}{form.cleaned_data['cuerpo_tlf']}"
            
            s.save()

            # --- REGISTRO EN BITÁCORA ---
            BitacoraAuditoria.objects.create(
                admin=request.user,
                accion='EDITAR_SOCIO',
                modulo='Socios',
                descripcion=(
                    f"Se actualizaron los datos del socio: {nombre_previo} (Unidad {unidad_previa}). "
                    f"Nuevos datos: {s.nombre} (Unidad {s.unidad})."
                ),
                ip_address=get_client_ip(request)
            )

            messages.success(request, "Datos actualizados correctamente.")
            return redirect('transparencia')
    else:
        form = SocioEditForm(instance=socio)

    context = {
        'form': form,
        'socio': socio,
        'acceso_activo': socio.user.is_active if socio.user else False
    }
    return render(request, 'core/editar_socio.html', context)

# --- ACCIONES DE SEGURIDAD ---

@user_passes_test(es_admin)
def toggle_acceso_portal(request, socio_id):
    """Activa o Bloquea el acceso al portal sin borrar al socio"""
    socio = get_object_or_404(Socio, pk=socio_id)

    if not socio.user:
        messages.error(request, "Este socio no tiene un usuario vinculado.")
        return redirect('editar_socio', socio_id)

    # Invertimos el estado (Si está activo lo desactiva, y viceversa)
    socio.user.is_active = not socio.user.is_active
    socio.user.save()

    estado = "HABILITADO" if socio.user.is_active else "BLOQUEADO"
    if socio.user.is_active:
        messages.success(request, f"Acceso {estado} para {socio.nombre}")
    else:
        messages.warning(request, f"Acceso {estado} para {socio.nombre}")

    return redirect('editar_socio', socio_id=socio_id)

@user_passes_test(es_admin)
def crear_usuario_portal(request, socio_id):
    """Crea un usuario automático si no lo tenía"""
    socio = get_object_or_404(Socio, pk=socio_id)

    if socio.user:
        messages.info(request, "El socio ya tiene usuario.")
        return redirect('editar_socio', socio_id)

    try:
        # Creamos usuario: Nombre = Unidad, Clave = Cedula
        username = socio.unidad.strip().lower()
        password = socio.cedula.strip()

        # Verificar si ya existe ese username
        if User.objects.filter(username=username).exists():
            username = f"{username}_{socio.id}" # Evitar duplicados

        user = User.objects.create_user(username=username, password=password, first_name=socio.nombre)
        socio.user = user
        socio.save()

        messages.success(request, f"Usuario creado. LOGIN: {username} | CLAVE: {password}")
    except Exception as e:
        messages.error(request, f"Error creando usuario: {e}")

    return redirect('editar_socio', socio_id)

@user_passes_test(es_admin)
def resetear_clave(request, socio_id):
    """Resetea la contraseña a la Cédula"""
    socio = get_object_or_404(Socio, pk=socio_id)

    if socio.user:
        nueva_clave = socio.cedula.strip()
        socio.user.set_password(nueva_clave)
        socio.user.save()
        messages.success(request, f"Clave restablecida. Nueva clave: {nueva_clave}")
    else:
        messages.error(request, "No tiene usuario.")

    return redirect('editar_socio', socio_id)

from django.db.models import Sum

@user_passes_test(es_admin)
def revision_transferencia_admin(request, pago_id):
    if not request.user.is_staff:
        return redirect('dashboard')
        
    pago_base = get_object_or_404(Pago, id=pago_id)
    referencia = pago_base.referencia
    socio_id = pago_base.aviso.socio.id
    
    pagos = Pago.objects.filter(
        referencia=referencia, 
        aviso__socio_id=socio_id
    ).select_related('aviso', 'metodo')
    
    totales = pagos.aggregate(
        usd=Sum('monto_dolares'),
        bs=Sum('monto_bolivares')
    )
    
    total_transferencia_usd = totales['usd'] or Decimal('0.00')
    total_transferencia_bs = totales['bs'] or Decimal('0.00')

    return render(request, 'core/revision_transferencia.html', {
        'pagos': pagos, 
        'socio': pago_base.aviso.socio, 
        'referencia': referencia,
        'metodo': pago_base.metodo, 
        'comprobante': pago_base.comprobante,
        'fecha_pago': pago_base.fecha_reporte, 
        'estado_global': pago_base.estado,
        'total_usd': total_transferencia_usd, 
        'total_bs': total_transferencia_bs,
        'pago_id_ejemplo': pago_base.id
    })

@login_required
def configuracion_perfil(request):
    socio = get_object_or_404(Socio, user=request.user)

    # Inicializamos los formularios
    form_tlf = ConfigTelefonoForm()
    form_pass = PasswordChangeForm(user=request.user)

    if request.method == 'POST':
        # ¿Viene de cambiar teléfono?
        if 'btn_telefono' in request.POST:
            form_tlf = ConfigTelefonoForm(request.POST)
            if form_tlf.is_valid():
                pref = form_tlf.cleaned_data['prefijo_tlf']
                cuerpo = form_tlf.cleaned_data['cuerpo_tlf']
                socio.telefono = f"58{pref[1:]}{cuerpo}"
                socio.save()
                messages.success(request, "✅ Teléfono actualizado correctamente.")
                return redirect('configuracion_perfil')

        # ¿Viene de cambiar contraseña?
        if 'btn_password' in request.POST:
            form_pass = PasswordChangeForm(user=request.user, data=request.POST)
            if form_pass.is_valid():
                user = form_pass.save()
                update_session_auth_hash(request, user)  # Mantiene la sesión iniciada
                messages.success(request, "🔐 Contraseña actualizada con éxito.")
                return redirect('configuracion_perfil')
            else:
                messages.error(request, "❌ Error al cambiar la contraseña. Revisa los datos.")

    # Pre-rellenar el teléfono actual si existe
    if socio.telefono and len(socio.telefono) == 12:
        # 58 414 1234567 -> extraemos 0414 y 1234567
        form_tlf.initial['prefijo_tlf'] = '0' + socio.telefono[2:5]
        form_tlf.initial['cuerpo_tlf'] = socio.telefono[5:]

    return render(request, 'core/config_perfil.html', {
        'form_tlf': form_tlf,
        'form_pass': form_pass,
        'socio': socio
    })

@user_passes_test(es_admin)
def gestionar_socio_cobranza(request, socio_id):
    socio = get_object_or_404(Socio, id=socio_id)
    
    # 1. Traemos los avisos usando prefetch para no saturar la base de datos
    avisos_pendientes = AvisoCobro.objects.filter(
        socio=socio
    ).exclude(estado='PAGADO').prefetch_related('detalles', 'pagos')

    # 2. Pagos que esperan tu aprobación
    pagos_raw = Pago.objects.filter(
        aviso__socio=socio, 
        estado='REVISION'
    ).select_related('aviso', 'metodo')

    transferencias = {}
    for p in pagos_raw:
        ref = p.referencia
        if ref not in transferencias:
            transferencias[ref] = {
                'pago_id': p.id,
                'referencia': ref,
                'monto_total_usd': Decimal('0.00'),
                'banco': p.metodo.nombre_banco if p.metodo else "Transferencia",
                'avisos_afectados': []
            }
        
        transferencias[ref]['monto_total_usd'] += p.monto_dolares
        # Usamos periodo_formateado que es más elegante que mes/anio
        periodo = p.aviso.periodo_formateado
        if periodo not in transferencias[ref]['avisos_afectados']:
            transferencias[ref]['avisos_afectados'].append(periodo)

    total_deuda = sum(aviso.saldo_pendiente_usd for aviso in avisos_pendientes)

    return render(request, 'core/gestionar_socio_cobranza.html', {
        'socio': socio, 
        'avisos': avisos_pendientes,
        'pagos_revision': transferencias.values(), 
        'total_deuda': total_deuda,
        'tasa': ConfigSistema.objects.first().tasa_bcv # La enviamos por si acaso
    })

from datetime import datetime, date
from decimal import Decimal
from django.db import transaction
from django.utils import timezone
import openpyxl

@login_required
@user_passes_test(lambda u: u.is_staff)
def importar_cargos_masivos(request):
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        excel_file = request.FILES['archivo_excel']
        
        # 1. Capturar el periodo desde el formulario
        try:
            mes_destino = int(request.POST.get('mes'))
            anio_destino = int(request.POST.get('anio'))
        except (TypeError, ValueError):
            messages.error(request, "Período inválido.")
            return redirect('importar_cargos_masivos')

        # 2. VALIDACIÓN DE CIERRE CONTABLE
        mes_cerrado = CierreMes.objects.filter(mes=mes_destino, anio=anio_destino).exists()
        
        if mes_cerrado:
            messages.error(
                request, 
                f"❌ ACCESO DENEGADO: El período {mes_destino}/{anio_destino} ya ha sido cerrado contablemente."
            )
            return redirect('modulo_cobranza')

        # 3. PROCESO DE IMPORTACIÓN
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            creados = 0
            errores = []

            with transaction.atomic():
                for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    unidad_val = str(row[0]).strip() if row[0] is not None else None
                    concepto_val = str(row[2]).strip() if row[2] is not None else None
                    monto_val = row[3]

                    if not unidad_val or not concepto_val or monto_val is None:
                        continue

                    socio = Socio.objects.filter(unidad=unidad_val).first()

                    if socio:
                        aviso, created = AvisoCobro.objects.get_or_create(
                            socio=socio,
                            mes=mes_destino,
                            anio=anio_destino,
                            defaults={
                                'fecha_emision': timezone.now().date(),
                                'estado': 'PENDIENTE'
                            }
                        )
                        
                        monto_limpio = Decimal(str(monto_val).replace(',', '.')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

                        ItemAviso.objects.create(
                            aviso=aviso,
                            descripcion=f"Importación: {concepto_val}",
                            monto_dolares=monto_limpio
                        )
                        
                        if aviso.estado == 'PAGADO':
                            aviso.estado = 'PARCIAL'
                            aviso.save()

                        creados += 1
                    else:
                        errores.append(f"Fila {index}: Unidad '{unidad_val}' no encontrada.")

            # --- REGISTRO EN BITÁCORA ---
            if creados > 0:
                BitacoraAuditoria.objects.create(
                    admin=request.user,
                    accion='IMPORTACION',
                    modulo='Cobranzas',
                    descripcion=(
                        f"Importación masiva de cargos desde Excel para el periodo {mes_destino}/{anio_destino}. "
                        f"Se procesaron {creados} cargos exitosamente. "
                        f"Archivo: {excel_file.name}. "
                        f"Errores en filas: {len(errores)}."
                    ),
                    ip_address=get_client_ip(request)
                )
                messages.success(request, f"✅ Importación finalizada: {creados} cargos añadidos.")
            
            if errores:
                for err in errores[:5]:
                    messages.warning(request, err)
                    
            return redirect('modulo_cobranza')

        except Exception as e:
            messages.error(request, f"❌ Error al procesar el archivo: {e}")

    return render(request, 'core/importar_cargos.html', {
        'meses': range(1, 13),
        'anios': range(2025, 2030)
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def cargar_deuda_manual_general(request):
    socios = Socio.objects.filter(activo=True).order_by('unidad')

    if request.method == 'POST':
        socio_id = request.POST.get('socio_id')
        descripcion = request.POST.get('descripcion')
        monto = request.POST.get('monto')
        fecha_manual = request.POST.get('fecha_cobro')

        if socio_id and descripcion and monto:
            socio = get_object_or_404(Socio, id=socio_id)

            # 1. Procesar la fecha
            if fecha_manual:
                fecha_final = datetime.strptime(fecha_manual, '%Y-%m-%d').date()
            else:
                fecha_final = timezone.now().date()

            mes_cobro = fecha_final.month
            anio_cobro = fecha_final.year

            # --- VALIDACIÓN DE CIERRE CONTABLE ---
            if CierreMes.objects.filter(mes=mes_cobro, anio=anio_cobro).exists():
                messages.error(
                    request, 
                    f"❌ No se puede cargar el cargo: El periodo {mes_cobro}/{anio_cobro} ya está cerrado."
                )
                return redirect('modulo_cobranza')
            # -------------------------------------

            # --- BARRERA DECIMAL ESTRICTA ---
            # Limpiamos posibles comas del formulario y forzamos a 2 decimales exactos
            monto_limpio = Decimal(str(monto).replace(',', '.')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

            with transaction.atomic():
                aviso, creado = AvisoCobro.objects.get_or_create(
                    socio=socio,
                    mes=mes_cobro,
                    anio=anio_cobro,
                    defaults={
                        'estado': 'PENDIENTE',
                        'fecha_emision': fecha_final
                    }
                )

                if not creado and aviso.estado == 'PAGADO':
                    aviso.estado = 'PARCIAL'
                    aviso.save()

                ItemAviso.objects.create(
                    aviso=aviso,
                    descripcion=f"Cargo Manual: {descripcion}",
                    monto_dolares=monto_limpio # Pasamos el Decimal limpio
                )

            BitacoraAuditoria.objects.create(
                admin=request.user,
                accion='CARGO_MANUAL',
                modulo='Cobranzas',
                descripcion=f"Cargo manual de ${monto_limpio} a la unidad {socio.unidad} para el periodo {mes_cobro}/{anio_cobro}. Motivo: {descripcion}",
                ip_address=get_client_ip(request)
            )

            # Usamos el monto_limpio en el mensaje para mostrar exactamente lo que se guardó
            messages.success(request, f"✅ Cargo de ${monto_limpio} agregado al Aviso {mes_cobro}/{anio_cobro}.")
            return redirect('modulo_cobranza')

    return render(request, 'core/form_cargo_manual_general.html', {'socios': socios})

@user_passes_test(es_admin)
def registrar_gasto_colectivo(request):
    config = ConfigSistema.objects.first()

    if request.method == 'POST':
        concepto = request.POST.get('concepto')
        monto_total_str = request.POST.get('monto')
        fecha_str = request.POST.get('fecha')
        archivo = request.FILES.get('comprobante')

        # 1. Procesar la fecha primero para validar el cierre
        try:
            fecha_final = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            fecha_final = timezone.now().date()
            
        mes_cobro = fecha_final.month
        anio_cobro = fecha_final.year

        # --- VALIDACIÓN DE CIERRE CONTABLE ---
        if CierreMes.objects.filter(mes=mes_cobro, anio=anio_cobro).exists():
            messages.error(
                request, 
                f"❌ Operación cancelada: El periodo {mes_cobro}/{anio_cobro} está cerrado contablemente."
            )
            return redirect('modulo_cobranza')
        # -------------------------------------

        socios_activos = Socio.objects.filter(activo=True)
        conteo = socios_activos.count()

        if conteo > 0 and monto_total_str and concepto:
            # --- BARRERA DECIMAL ESTRICTA ---
            monto_total_limpio = Decimal(monto_total_str.replace(',', '.')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            monto_por_socio = (monto_total_limpio / Decimal(conteo)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            
            try:
                with transaction.atomic():
                    # Registramos el Gasto Global
                    GastoGeneral.objects.create(
                        concepto=concepto,
                        monto_total_usd=monto_total_limpio,
                        fecha_factura=fecha_final,
                        comprobante=archivo
                    )

                    # Repartimos la cuota
                    for socio in socios_activos:
                        aviso, creado = AvisoCobro.objects.get_or_create(
                            socio=socio,
                            mes=mes_cobro,
                            anio=anio_cobro,
                            defaults={
                                'estado': 'PENDIENTE',
                                'fecha_emision': fecha_final
                            }
                        )

                        if not creado and aviso.estado == 'PAGADO':
                            aviso.estado = 'PARCIAL'
                            aviso.save()

                        ItemAviso.objects.create(
                            aviso=aviso,
                            descripcion=f"Cuota proporcional de: {concepto}",
                            monto_dolares=monto_por_socio
                        )

                # --- REGISTRO EN BITÁCORA (Fuera del atomic para asegurar éxito) ---
                BitacoraAuditoria.objects.create(
                    admin=request.user,
                    accion='GASTO_COLECTIVO', # Recuerda agregar esto a los CHOICES de tu modelo si no existe
                    modulo='Cobranzas',
                    descripcion=(
                        f"Distribución de gasto colectivo: '{concepto}'. "
                        f"Monto total: ${monto_total_limpio}. "
                        f"Monto asignado por socio: ${monto_por_socio}. "
                        f"Total socios afectados: {conteo}."
                    ),
                    ip_address=get_client_ip(request)
                )

                messages.success(request, f"✅ Gasto '{concepto}' distribuido correctamente.")
                return redirect('modulo_cobranza')

            except Exception as e:
                messages.error(request, f"❌ Error al procesar el gasto colectivo: {e}")

    context = {
        'config': config,
        'fecha_actual': timezone.now(),
    }
    return render(request, 'core/form_gasto_comun.html', context)

def obtener_o_crear_aviso_actual(socio):
    hoy = timezone.now()
    # Busca el aviso de este mes y año. Si no existe, lo crea.
    aviso, creado = AvisoCobro.objects.get_or_create(
        socio=socio,
        mes=hoy.month,
        anio=hoy.year,
        defaults={'estado': 'PENDIENTE'}
    )
    return aviso

def cargar_deuda_a_socio(socio, descripcion, monto_usd):
    """Agrega un nuevo ítem al aviso de cobro del mes actual del socio."""
    
    # 1. Agarramos el "papel" (Aviso) de este mes
    aviso = obtener_o_crear_aviso_actual(socio)
    
    # 2. Le escribimos la nueva "línea" (Item)
    nuevo_item = ItemAviso.objects.create(
        aviso=aviso,
        descripcion=descripcion,
        monto_dolares=monto_usd
    )
    
    # 3. LÓGICA DE NEGOCIO CLAVE:
    # Si el socio ya había pagado su mes, pero le metes una multa nueva, 
    # el aviso debe volver a estar 'PENDIENTE' (o 'PARCIAL' si ya había plata abonada).
    if aviso.estado == 'PAGADO':
        aviso.estado = 'PARCIAL'
        aviso.save()
        
    return nuevo_item


@user_passes_test(es_admin)
def importar_socios_masivo(request):
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        excel_file = request.FILES['archivo_excel']
        PREFIJO_LINEA = "unidos"
        try:
            # Abrimos en modo lectura para ser eficientes con la memoria
            wb = openpyxl.load_workbook(excel_file, read_only=True)
            sheet = wb.active
            creados = 0
            errores = []

            # Recorremos el Excel fila por fila
            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Columnas esperadas: 0:Nombre, 1:Unidad, 2:Nacionalidad, 3:Cedula, 4:Telefono
                nombre_raw = row[0]
                unidad_raw = row[1]
                nacionalidad_raw = row[2]
                cedula_raw = row[3]
                tlf_raw = row[4]

                # Validación básica de campos obligatorios
                if not (nombre_raw and unidad_raw and cedula_raw):
                    continue

                try:
                    # Usamos transacciones individuales por socio para no bloquear SQLite
                    with transaction.atomic():
                        # 1. Formatear Unidad (ej: de "1" a "001")
                        unidad_limpia = str(unidad_raw).strip().zfill(3)
                        
                        # 2. Generar Username (ej: unidos001)
                        username = f"{PREFIJO_LINEA}{unidad_limpia}".lower()

                        # 3. Validar duplicados
                        if User.objects.filter(username=username).exists():
                            errores.append(f"Fila {index}: El usuario '{username}' ya existe.")
                            continue
                        
                        if Socio.objects.filter(unidad=unidad_limpia).exists():
                            errores.append(f"Fila {index}: La unidad '{unidad_limpia}' ya está registrada.")
                            continue

                        # 4. Formatear Cédula (V012345678)
                        nac = str(nacionalidad_raw).strip().upper() if nacionalidad_raw else 'V'
                        cedula_f = f"{nac}{str(cedula_raw).strip().zfill(9)}"

                        # 5. Formatear Teléfono (58414...)
                        tlf_str = str(tlf_raw).strip()
                        if tlf_str.startswith('0'):
                            telefono_f = f"58{tlf_str[1:]}"
                        elif not tlf_str.startswith('58') and len(tlf_str) > 0:
                            telefono_f = f"58{tlf_str}"
                        else:
                            telefono_f = tlf_str

                        # 6. Crear Usuario de Django (Clave = Cédula)
                        user = User.objects.create_user(
                            username=username,
                            password=cedula_f,
                            first_name=str(nombre_raw).strip()[:150] # Límite de Django
                        )

                        # 7. Crear Perfil de Socio
                        Socio.objects.create(
                            user=user,
                            nombre=str(nombre_raw).strip(),
                            unidad=unidad_limpia,
                            cedula=cedula_f,
                            telefono=telefono_f,
                            activo=True
                        )
                        creados += 1

                except Exception as e:
                    errores.append(f"Fila {index}: Error al crear ({e})")

            # Feedback al usuario
            if creados > 0:
                resumen_errores = f" con {len(errores)} errores encontrados" if errores else " sin errores"
                
                BitacoraAuditoria.objects.create(
                    admin=request.user,
                    accion='IMPORTACION',
                    modulo='Socios',
                    descripcion=f"Importación masiva de socios. Creados: {creados} registros{resumen_errores}. Archivo: {excel_file.name}",
                    ip_address=get_client_ip(request)
                )

                messages.success(request, f"✅ Se importaron {creados} socios exitosamente con el formato '{PREFIJO_LINEA}###'.")
            
            if errores:
                for err in errores[:5]: # Mostramos los primeros 5 para no saturar
                    messages.warning(request, err)
                if len(errores) > 5:
                    messages.info(request, f"...y otros {len(errores)-5} errores más.")

            return redirect('transparencia')

        except Exception as e:
            messages.error(request, f"❌ Error crítico procesando el archivo: {e}")

    return render(request, 'core/importar_socios.html')

@user_passes_test(es_admin)
def descargar_plantilla_socios(request):
    # Crear el libro de trabajo y la hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla de Socios"

    # 1. Definir Encabezados
    headers = ['Nombre Completo', 'Nro Socio', 'Nacionalidad (V/E)', 'Cédula (Solo números)', 'Teléfono (04xxxxxxxx)']
    ws.append(headers)

    # 2. Estilo para los encabezados (Negrita, color y alineación)
    header_fill = PatternFill(start_color="343a40", end_color="343a40", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 3. Agregar un ejemplo (Opcional, ayuda al usuario)
    ws.append(['Arturo Acevedo', '001', 'V', '12345678', '04141234567'])
    
    # Ajustar el ancho de las columnas para que se vea bien
    column_widths = [30, 10, 20, 20, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64+i)].width = width

    # 4. Preparar la respuesta del navegador
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=plantilla_socios_unidos.xlsx'
    
    wb.save(response)
    return response

from django.contrib.auth.forms import SetPasswordForm

@login_required
def cambiar_clave_inicial(request):
    try:
        socio = request.user.socio
    except AttributeError:
        # Si un admin entra aquí por error, lo mandamos al dashboard
        return redirect('dashboard')

    if request.method == 'POST':
        # Usamos SetPasswordForm porque NO pide la clave anterior
        form = SetPasswordForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            socio.clave_cambiada = True
            socio.save()
            messages.success(request, "¡Contraseña actualizada! Ya puedes usar el portal.")
            return redirect('portal_socio')
    else:
        form = SetPasswordForm(request.user)
    
    return render(request, 'core/cambiar_clave_obligatorio.html', {'form': form})

@login_required
@user_passes_test(lambda u: u.is_staff)
def eliminar_metodo_pago(request, metodo_id):
    metodo = get_object_or_404(MetodoPago, id=metodo_id)
    
    # Capturamos los datos antes de borrar el objeto de la BD
    tipo = metodo.get_tipo_display()
    banco = metodo.nombre_banco
    nombre_cuenta = metodo.nombre
    
    # 1. Ejecutamos la eliminación
    metodo.delete()
    
    # 2. REGISTRO EN BITÁCORA (El "Quién hizo qué")
    BitacoraAuditoria.objects.create(
        admin=request.user,
        accion='ELIMINAR_METODO', # Asegúrate de que coincida con tus CHOICES
        modulo='Configuración',
        descripcion=f"Se eliminó permanentemente el método de pago: {tipo} - {banco} ({nombre_cuenta})",
        ip_address=get_client_ip(request)
    )
    
    messages.success(request, f"El método {tipo} de {banco} ha sido eliminado permanentemente.")
    return redirect('configuracion_pagos')

@user_passes_test(es_admin)
def cerrar_mes_contable(request):
    if request.method == 'POST':
        # 1. Convertimos a int para asegurar consistencia con el modelo
        try:
            mes = int(request.POST.get('mes'))
            anio = int(request.POST.get('anio'))
        except (ValueError, TypeError):
            messages.error(request, "Formato de mes o año inválido.")
            return redirect('dashboard')
        
        cierre, creado = CierreMes.objects.get_or_create(
            mes=mes, 
            anio=anio, 
            defaults={'cerrado_por': request.user}
        )
        
        if creado:
            # TRACKING DE AUDITORÍA: Ahora sí se ejecutará al crear el registro
            BitacoraAuditoria.objects.create(
                admin=request.user,
                accion='CIERRE_MES',
                modulo='Contabilidad',
                descripcion=f"Cierre de periodo contable definitivo para el mes {mes} del año {anio}.",
                ip_address=get_client_ip(request)
            )
            messages.success(request, f"✅ Periodo {mes}/{anio} cerrado exitosamente.")
        else:
            # Si entras aquí, es que 'creado' es False, por eso no veías nada en la bitácora
            messages.warning(request, f"El periodo {mes}/{anio} ya se encuentra cerrado en el sistema.")
            
    return redirect('dashboard')
@user_passes_test(es_admin)
def gestion_periodos(request):
    hoy = timezone.now()
    
    # 1. Inteligencia de Años (lo que ya teníamos)
    primer_aviso = AvisoCobro.objects.order_by('anio').first()
    primer_cierre = CierreMes.objects.order_by('anio').first()
    anio_min = hoy.year
    if primer_aviso: anio_min = min(anio_min, primer_aviso.anio)
    if primer_cierre: anio_min = min(anio_min, primer_cierre.anio)
    
    anios_disponibles = sorted(range(anio_min, hoy.year + 1), reverse=True)
    anio_seleccionado = int(request.GET.get('anio', hoy.year))

    # 2. CÁLCULOS DEL REPORTE ANUAL
    # Filtramos avisos del año seleccionado
    avisos_del_anio = AvisoCobro.objects.filter(anio=anio_seleccionado)
    
    # Total Facturado (Suma de todos los ítems de esos avisos)
    total_facturado_anio = ItemAviso.objects.filter(
        aviso__anio=anio_seleccionado
    ).aggregate(total=Sum('monto_dolares'))['total'] or Decimal('0.00')

    # Total Recaudado (Suma de todos los pagos APROBADOS de esos avisos)
    total_recaudado_anio = Pago.objects.filter(
        aviso__anio=anio_seleccionado, 
        estado='APROBADO'
    ).aggregate(total=Sum('monto_dolares'))['total'] or Decimal('0.00')

    # Cálculos derivados
    por_cobrar_anio = total_facturado_anio - total_recaudado_anio
    porcentaje_eficiencia = (
        int((total_recaudado_anio / total_facturado_anio) * 100) 
        if total_facturado_anio > 0 else 0
    )

    # 3. LISTADO DE MESES (lo que ya teníamos)
    periodos_cerrados = set(CierreMes.objects.filter(anio=anio_seleccionado).values_list('mes', flat=True))
    meses_nombres = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]

    meses_info = []
    for num, nombre in meses_nombres:
        esta_cerrado = num in periodos_cerrados
        conteo_avisos = AvisoCobro.objects.filter(mes=num, anio=anio_seleccionado).count()
        meses_info.append({
            'numero': num, 'nombre': nombre, 'cerrado': esta_cerrado, 'avisos': conteo_avisos
        })

    return render(request, 'core/gestion_periodos.html', {
        'anio_seleccionado': anio_seleccionado,
        'anios_disponibles': anios_disponibles,
        'meses': meses_info,
        'resumen': {
            'facturado': total_facturado_anio,
            'recaudado': total_recaudado_anio,
            'pendiente': por_cobrar_anio,
            'eficiencia': porcentaje_eficiencia
        }
    })

@user_passes_test(es_admin)
def bitacora_auditoria(request):
    # 1. Base del QuerySet
    registros_list = BitacoraAuditoria.objects.select_related('admin').all().order_by('-fecha')

    # 2. Captura de filtros
    f_accion = request.GET.get('accion')
    f_admin = request.GET.get('admin')
    f_desde = request.GET.get('desde')
    f_hasta = request.GET.get('hasta')

    # 3. Aplicación de filtros
    if f_accion:
        registros_list = registros_list.filter(accion=f_accion)
    if f_admin:
        registros_list = registros_list.filter(admin_id=f_admin)
    if f_desde:
        registros_list = registros_list.filter(fecha__date__gte=f_desde)
    if f_hasta:
        registros_list = registros_list.filter(fecha__date__lte=f_hasta)

    # Lógica de Exportación (Manteniendo los filtros aplicados)
     # --- LÓGICA DE EXPORTACIÓN ---

    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="auditoria_conductores_unidos.csv"'
        writer = csv.writer(response)
        writer.writerow(['Fecha', 'Administrador', 'Acción', 'Módulo', 'Descripción', 'IP Address'])

        for log in registros_list:
            writer.writerow([
                log.fecha.strftime('%d/%m/%Y %H:%M'),
                log.admin.username if log.admin else 'Sistema',
                log.get_accion_display(),
                log.modulo,
                log.descripcion,
                log.ip_address
            ])

        return response

    # -----------------------------

    # Paginación
    paginator = Paginator(registros_list, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj,
        'admins': User.objects.filter(is_staff=True),
        'acciones_list': BitacoraAuditoria.ACCIONES, # Pasamos la tupla de acciones
        # Devolvemos los filtros para mantenerlos en los inputs
        'f_accion': f_accion,
        'f_admin': f_admin,
        'f_desde': f_desde,
        'f_hasta': f_hasta,
    }
    
    return render(request, 'core/auditoria.html', context)

def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def validar_solvencia(request, slug_verificacion):
    socio = get_object_or_404(Socio, codigo_verificacion=slug_verificacion)
    
    solvente = socio.es_solvente
    
    context = {
        'socio': socio,
        'solvente': solvente,
        'fecha': timezone.now(),
    }
    # Usaremos un template "limpio" diseñado para móviles
    return render(request, 'core/validar_solvencia.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff)
def reporte_financiero_conceptos(request):
    hoy = timezone.now()
    anio_sel = int(request.GET.get('anio', hoy.year))
    mes_sel = int(request.GET.get('mes', hoy.month))

    # 1. Obtener años disponibles para el filtro
    anios_disponibles = AvisoCobro.objects.values_list('anio', flat=True).distinct().order_by('-anio')
    if not anios_disponibles:
        anios_disponibles = [hoy.year]

    # 2. Consultar y agrupar los ítems facturados en ese periodo
    # Agrupamos por 'descripcion' y sumamos los montos
    desglose_conceptos = ItemAviso.objects.filter(
        aviso__anio=anio_sel, 
        aviso__mes=mes_sel
    ).values('descripcion').annotate(
        total_usd=Sum('monto_dolares'),
        cantidad_cargos=Count('id')
    ).order_by('-total_usd')

    # Total general del mes
    total_mes_usd = sum((item['total_usd'] for item in desglose_conceptos), Decimal('0.00'))

    # --- LÓGICA DE EXPORTACIÓN A EXCEL ---
    if request.GET.get('export') == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = f"Reporte_{mes_sel}_{anio_sel}"

        # Estilos de encabezado
        header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        ws.append(['Concepto Billed', 'Cantidad de Cargos Generados', 'Total Proyectado (USD)'])
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Llenar datos
        for item in desglose_conceptos:
            ws.append([
                item['descripcion'], 
                item['cantidad_cargos'], 
                float(item['total_usd']) # Excel maneja mejor floats que Decimals
            ])
            
        # Fila de totales
        ws.append(['', 'TOTAL FACTURADO:', float(total_mes_usd)])
        ws[ws.max_row][1].font = Font(bold=True)
        ws[ws.max_row][2].font = Font(bold=True)

        # Ajustar anchos
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 25

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Desglose_Conceptos_{mes_sel}_{anio_sel}.xlsx'
        wb.save(response)
        return response

    # Nombres de meses para el selector HTML
    meses_nombres = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]

    context = {
        'desglose': desglose_conceptos,
        'total_mes_usd': total_mes_usd,
        'anio_sel': anio_sel,
        'mes_sel': mes_sel,
        'anios_disponibles': anios_disponibles,
        'meses_nombres': meses_nombres,
    }
    return render(request, 'core/reporte_conceptos.html', context)